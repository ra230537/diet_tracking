"""
Diet Plan Router
=================
Endpoints for managing diet plans, variations, meals, and meal items.

Endpoints:
  POST   /diet/plans                          - Create a new diet plan
  GET    /diet/current                        - Get the full current diet plan with calculated macros
  PUT    /diet/plans/{id}/targets             - Update diet plan macro targets
  POST   /diet/plans/{id}/variations          - Create a new variation (empty or duplicated)
  PATCH  /diet/variations/{id}                - Rename a variation
  DELETE /diet/variations/{id}                - Delete a variation and all its meals
  POST   /diet/variations/{id}/meals          - Add a meal to a variation
  POST   /diet/plans/{id}/meals               - Add a meal to a plan (backward compat)
  PATCH  /diet/meals/{id}                     - Rename a meal
  DELETE /diet/meals/{id}                     - Delete a meal and all its items
  POST   /diet/meals/{id}/add_item            - Add a food item to a meal
  PUT    /diet/meal-items/{id}                - Update meal item quantity
  DELETE /diet/meal-items/{id}                - Remove a food item from a meal
  GET    /diet/export/excel                   - Export diet as Excel
  GET    /diet/export/pdf                     - Export diet as PDF
"""

import io
import logging

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.models import DietPlan, DietVariation, FoodItem, Meal, MealItem
from app.schemas import (
    DietPlanCreate,
    DietPlanFullResponse,
    DietPlanResponse,
    DietPlanUpdate,
    DietVariationCreate,
    DietVariationRename,
    DietVariationResponse,
    MealCreate,
    MealItemCreate,
    MealItemResponse,
    MealItemUpdate,
    MealRename,
    MealResponse,
    MessageResponse,
)
from app.services.diet_calculator import get_current_diet_full

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/diet", tags=["Diet Plans"])


@router.post("/plans", response_model=DietPlanResponse, status_code=201)
async def create_diet_plan(
    plan: DietPlanCreate,
    db: AsyncSession = Depends(get_db),
):
    """
    Create a new diet plan with daily macro targets.
    
    If is_active=True (default), any other active plans for this user
    will be deactivated automatically. Only one plan can be active at a time.
    
    Example request:
    {
        "target_calories": 3000,
        "target_protein": 180,
        "target_carbs": 375,
        "target_fat": 83
    }
    """
    # If the new plan should be active, deactivate all other plans for this user
    if plan.is_active:
        stmt = (
            select(DietPlan)
            .where(DietPlan.user_id == plan.user_id)
            .where(DietPlan.is_active == True)
        )
        result = await db.execute(stmt)
        existing_active_plans = result.scalars().all()

        for existing_plan in existing_active_plans:
            existing_plan.is_active = False
            logger.info(f"Deactivated plan ID {existing_plan.id} for user '{plan.user_id}'")

    # Create the new plan
    db_plan = DietPlan(**plan.model_dump())
    db.add(db_plan)
    await db.flush()  # Get the plan ID

    # Automatically create a default "Principal" variation
    default_variation = DietVariation(
        diet_plan_id=db_plan.id,
        name="Principal",
        order_index=0,
    )
    db.add(default_variation)

    await db.commit()
    await db.refresh(db_plan)

    logger.info(
        f"Created diet plan ID {db_plan.id} for user '{db_plan.user_id}' "
        f"(calories={db_plan.target_calories}, protein={db_plan.target_protein}g)"
    )
    return db_plan


@router.get("/current", response_model=DietPlanFullResponse)
async def get_current_diet(
    user_id: str = "default_user",
    db: AsyncSession = Depends(get_db),
):
    """
    Get the complete current (active) diet plan with full hierarchy.
    
    Returns:
      - Plan metadata (targets)
      - All meals with their items
      - Calculated macros for each item (based on quantity * per-100g values)
      - Meal-level totals
      - Plan-level grand totals
      - Target vs Actual comparison for each macro
    
    This is the primary endpoint for the diet overview dashboard.
    """
    try:
        plan_data = await get_current_diet_full(db, user_id)
        return plan_data
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


# ============================================================
# VARIATION ENDPOINTS
# ============================================================

@router.post("/plans/{plan_id}/variations", response_model=DietVariationResponse, status_code=201)
async def create_variation(
    plan_id: int,
    variation: DietVariationCreate,
    duplicate_from: int | None = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Create a new variation for a diet plan.

    Options:
      - Create from scratch: just provide name and order_index
      - Duplicate from existing: pass ?duplicate_from=<variation_id> to copy all meals and items

    Example (from scratch):
      POST /diet/plans/1/variations
      { "name": "Substituição", "order_index": 1 }

    Example (duplicate):
      POST /diet/plans/1/variations?duplicate_from=5
      { "name": "Substituição (cópia)", "order_index": 1 }
    """
    # Verify the plan exists
    stmt = select(DietPlan).where(DietPlan.id == plan_id)
    result = await db.execute(stmt)
    plan = result.scalar_one_or_none()

    if not plan:
        raise HTTPException(status_code=404, detail=f"Diet plan with ID {plan_id} not found.")

    # Create the variation
    db_variation = DietVariation(
        diet_plan_id=plan_id,
        name=variation.name,
        order_index=variation.order_index,
    )
    db.add(db_variation)
    await db.flush()  # Get the ID without committing

    meals_data: list[dict] = []

    # If duplicating from an existing variation
    if duplicate_from is not None:
        source_stmt = (
            select(DietVariation)
            .where(DietVariation.id == duplicate_from)
            .where(DietVariation.diet_plan_id == plan_id)
            .options(
                selectinload(DietVariation.meals)
                .selectinload(Meal.items)
                .selectinload(MealItem.food_item)
            )
        )
        source_result = await db.execute(source_stmt)
        source_variation = source_result.scalar_one_or_none()

        if not source_variation:
            raise HTTPException(
                status_code=404,
                detail=f"Source variation with ID {duplicate_from} not found in plan {plan_id}."
            )

        # Copy meals and items
        for source_meal in source_variation.meals:
            new_meal = Meal(
                diet_plan_id=plan_id,
                variation_id=db_variation.id,
                name=source_meal.name,
                order_index=source_meal.order_index,
            )
            db.add(new_meal)
            await db.flush()

            items_data = []
            meal_total_cal = meal_total_pro = meal_total_carb = meal_total_fat = 0.0

            for source_item in source_meal.items:
                new_item = MealItem(
                    meal_id=new_meal.id,
                    food_item_id=source_item.food_item_id,
                    quantity_grams=source_item.quantity_grams,
                )
                db.add(new_item)
                await db.flush()

                qty_factor = source_item.quantity_grams / 100
                cal = round(qty_factor * source_item.food_item.calories_kcal, 2)
                pro = round(qty_factor * source_item.food_item.protein_g, 2)
                carb = round(qty_factor * source_item.food_item.carbs_g, 2)
                fat = round(qty_factor * source_item.food_item.fat_g, 2)

                meal_total_cal += cal
                meal_total_pro += pro
                meal_total_carb += carb
                meal_total_fat += fat

                items_data.append(MealItemResponse(
                    id=new_item.id,
                    food_item_id=source_item.food_item_id,
                    food_item_name=source_item.food_item.name,
                    quantity_grams=source_item.quantity_grams,
                    calculated_calories=cal,
                    calculated_protein=pro,
                    calculated_carbs=carb,
                    calculated_fat=fat,
                ))

            meals_data.append({
                "id": new_meal.id,
                "name": new_meal.name,
                "order_index": new_meal.order_index,
                "items": items_data,
                "total_calories": round(meal_total_cal, 2),
                "total_protein": round(meal_total_pro, 2),
                "total_carbs": round(meal_total_carb, 2),
                "total_fat": round(meal_total_fat, 2),
            })

    await db.commit()
    await db.refresh(db_variation)

    logger.info(f"Created variation '{db_variation.name}' for plan ID {plan_id}"
                f"{' (duplicated from ' + str(duplicate_from) + ')' if duplicate_from else ''}")

    # Build response
    variation_total_cal = sum(m.get("total_calories", 0) for m in meals_data)
    variation_total_pro = sum(m.get("total_protein", 0) for m in meals_data)
    variation_total_carb = sum(m.get("total_carbs", 0) for m in meals_data)
    variation_total_fat = sum(m.get("total_fat", 0) for m in meals_data)

    return DietVariationResponse(
        id=db_variation.id,
        name=db_variation.name,
        order_index=db_variation.order_index,
        created_at=db_variation.created_at,
        meals=[MealResponse(**m) if isinstance(m, dict) else m for m in meals_data],
        total_calories=round(variation_total_cal, 2),
        total_protein=round(variation_total_pro, 2),
        total_carbs=round(variation_total_carb, 2),
        total_fat=round(variation_total_fat, 2),
    )


@router.patch("/variations/{variation_id}", response_model=DietVariationResponse)
async def rename_variation(
    variation_id: int,
    payload: DietVariationRename,
    user_id: str = "default_user",
    db: AsyncSession = Depends(get_db),
):
    """Rename an existing variation."""
    stmt = (
        select(DietVariation)
        .where(DietVariation.id == variation_id)
        .options(
            selectinload(DietVariation.diet_plan),
            selectinload(DietVariation.meals)
            .selectinload(Meal.items)
            .selectinload(MealItem.food_item),
        )
    )
    result = await db.execute(stmt)
    variation = result.scalar_one_or_none()

    if not variation:
        raise HTTPException(status_code=404, detail=f"Variation with ID {variation_id} not found.")

    if variation.diet_plan.user_id != user_id:
        raise HTTPException(status_code=403, detail="This variation does not belong to your diet plan.")

    old_name = variation.name
    variation.name = payload.name
    await db.commit()
    await db.refresh(variation)

    logger.info(f"Renamed variation ID {variation_id} from '{old_name}' to '{payload.name}'")

    # Build response with calculated totals
    from app.services.diet_calculator import _build_variation_data
    var_data = _build_variation_data(variation)

    return DietVariationResponse(**var_data)


@router.delete("/variations/{variation_id}", response_model=MessageResponse)
async def delete_variation(
    variation_id: int,
    user_id: str = "default_user",
    db: AsyncSession = Depends(get_db),
):
    """Delete a variation and all its meals/items. Cannot delete the last variation."""
    stmt = (
        select(DietVariation)
        .where(DietVariation.id == variation_id)
        .options(selectinload(DietVariation.diet_plan).selectinload(DietPlan.variations))
    )
    result = await db.execute(stmt)
    variation = result.scalar_one_or_none()

    if not variation:
        raise HTTPException(status_code=404, detail=f"Variation with ID {variation_id} not found.")

    if variation.diet_plan.user_id != user_id:
        raise HTTPException(status_code=403, detail="This variation does not belong to your diet plan.")

    # Don't allow deleting the last variation
    if len(variation.diet_plan.variations) <= 1:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete the only remaining variation. A plan must have at least one variation."
        )

    var_name = variation.name
    await db.delete(variation)
    await db.commit()

    logger.info(f"Deleted variation '{var_name}' (ID {variation_id}) and all its meals")

    return MessageResponse(
        message="Variação excluída com sucesso.",
        detail=f"Variação '{var_name}' (ID {variation_id}) e todas as refeições associadas foram removidas."
    )


@router.post("/variations/{variation_id}/meals", response_model=MealResponse, status_code=201)
async def add_meal_to_variation(
    variation_id: int,
    meal: MealCreate,
    db: AsyncSession = Depends(get_db),
):
    """
    Add a new meal to a specific variation.
    """
    stmt = (
        select(DietVariation)
        .where(DietVariation.id == variation_id)
        .options(selectinload(DietVariation.meals))
    )
    result = await db.execute(stmt)
    variation = result.scalar_one_or_none()

    if not variation:
        raise HTTPException(status_code=404, detail=f"Variation with ID {variation_id} not found.")

    db_meal = Meal(
        diet_plan_id=variation.diet_plan_id,
        variation_id=variation_id,
        name=meal.name,
        order_index=meal.order_index,
    )
    db.add(db_meal)
    await db.commit()
    await db.refresh(db_meal)

    logger.info(f"Added meal '{db_meal.name}' to variation ID {variation_id}")

    return MealResponse(
        id=db_meal.id,
        name=db_meal.name,
        order_index=db_meal.order_index,
        items=[],
        total_calories=0.0,
        total_protein=0.0,
        total_carbs=0.0,
        total_fat=0.0,
    )


# ============================================================
# BACKWARD-COMPATIBLE MEAL ENDPOINTS
# ============================================================

@router.post("/plans/{plan_id}/meals", response_model=MealResponse, status_code=201)
async def add_meal_to_plan(
    plan_id: int,
    meal: MealCreate,
    db: AsyncSession = Depends(get_db),
):
    """
    Add a new meal to an existing diet plan.
    If the plan has variations, the meal is added to the first variation.
    
    Meals represent eating occasions during the day (e.g., "Breakfast",
    "Post-workout shake", "Dinner"). The order_index controls display order.
    """
    # Verify the plan exists with its variations
    stmt = (
        select(DietPlan)
        .where(DietPlan.id == plan_id)
        .options(selectinload(DietPlan.variations))
    )
    result = await db.execute(stmt)
    plan = result.scalar_one_or_none()

    if not plan:
        raise HTTPException(
            status_code=404,
            detail=f"Diet plan with ID {plan_id} not found."
        )

    # Get the first variation (or None)
    variation_id = None
    if plan.variations:
        first_variation = sorted(plan.variations, key=lambda v: v.order_index)[0]
        variation_id = first_variation.id

    # Create the meal
    db_meal = Meal(
        diet_plan_id=plan_id,
        variation_id=variation_id,
        name=meal.name,
        order_index=meal.order_index,
    )
    db.add(db_meal)
    await db.commit()
    await db.refresh(db_meal)

    logger.info(f"Added meal '{db_meal.name}' to plan ID {plan_id}")

    return MealResponse(
        id=db_meal.id,
        name=db_meal.name,
        order_index=db_meal.order_index,
        items=[],
        total_calories=0.0,
        total_protein=0.0,
        total_carbs=0.0,
        total_fat=0.0,
    )


@router.patch("/meals/{meal_id}", response_model=MealResponse)
async def rename_meal(
    meal_id: int,
    payload: MealRename,
    user_id: str = "default_user",
    db: AsyncSession = Depends(get_db),
):
    """
    Rename an existing meal within the user's active diet plan.

    Verifies that the meal belongs to the authenticated user's active plan
    before applying the update.

    Example:
      PATCH /diet/meals/3
      { "name": "Post-workout Shake" }
    """
    # Fetch the meal with its parent plan loaded
    stmt = (
        select(Meal)
        .where(Meal.id == meal_id)
        .options(selectinload(Meal.diet_plan), selectinload(Meal.items).selectinload(MealItem.food_item))
    )
    result = await db.execute(stmt)
    meal = result.scalar_one_or_none()

    if not meal:
        raise HTTPException(status_code=404, detail=f"Meal with ID {meal_id} not found.")

    # Verify ownership: meal must belong to this user's active plan
    if meal.diet_plan.user_id != user_id or not meal.diet_plan.is_active:
        raise HTTPException(
            status_code=403,
            detail="This meal does not belong to your active diet plan."
        )

    old_name = meal.name
    meal.name = payload.name
    await db.commit()
    await db.refresh(meal)

    logger.info(f"Renamed meal ID {meal_id} from '{old_name}' to '{payload.name}'")

    # Build response with totals
    items_response = []
    total_cal = total_prot = total_carbs = total_fat = 0.0
    for item in meal.items:
        qty_factor = item.quantity_grams / 100
        cal = round(qty_factor * item.food_item.calories_kcal, 2)
        prot = round(qty_factor * item.food_item.protein_g, 2)
        carbs = round(qty_factor * item.food_item.carbs_g, 2)
        fat = round(qty_factor * item.food_item.fat_g, 2)
        total_cal += cal
        total_prot += prot
        total_carbs += carbs
        total_fat += fat
        items_response.append(MealItemResponse(
            id=item.id,
            food_item_id=item.food_item.id,
            food_item_name=item.food_item.name,
            quantity_grams=item.quantity_grams,
            calculated_calories=cal,
            calculated_protein=prot,
            calculated_carbs=carbs,
            calculated_fat=fat,
        ))

    return MealResponse(
        id=meal.id,
        name=meal.name,
        order_index=meal.order_index,
        items=items_response,
        total_calories=round(total_cal, 2),
        total_protein=round(total_prot, 2),
        total_carbs=round(total_carbs, 2),
        total_fat=round(total_fat, 2),
    )


@router.delete("/meals/{meal_id}", response_model=MessageResponse)
async def delete_meal(
    meal_id: int,
    user_id: str = "default_user",
    db: AsyncSession = Depends(get_db),
):
    """
    Delete a meal and all its associated meal items (cascade).

    Verifies that the meal belongs to the authenticated user's active plan
    before deleting.

    The SQLAlchemy relationship cascade="all, delete-orphan" on Meal.items
    ensures all MealItems are automatically removed.

    Example:
      DELETE /diet/meals/3?user_id=default_user
    """
    # Fetch the meal with its parent plan loaded
    stmt = (
        select(Meal)
        .where(Meal.id == meal_id)
        .options(selectinload(Meal.diet_plan))
    )
    result = await db.execute(stmt)
    meal = result.scalar_one_or_none()

    if not meal:
        raise HTTPException(status_code=404, detail=f"Meal with ID {meal_id} not found.")

    # Verify ownership: meal must belong to this user's active plan
    if meal.diet_plan.user_id != user_id or not meal.diet_plan.is_active:
        raise HTTPException(
            status_code=403,
            detail="This meal does not belong to your active diet plan."
        )

    meal_name = meal.name
    await db.delete(meal)
    await db.commit()

    logger.info(f"Deleted meal '{meal_name}' (ID {meal_id}) and all its items")

    return MessageResponse(
        message="Meal deleted successfully.",
        detail=f"Removed meal '{meal_name}' (ID {meal_id}) and all associated items."
    )


@router.post("/meals/{meal_id}/add_item", response_model=MealItemResponse, status_code=201)
async def add_item_to_meal(
    meal_id: int,
    item: MealItemCreate,
    db: AsyncSession = Depends(get_db),
):
    """
    Add a food item to a meal with a specific quantity.
    
    The quantity_grams specifies how much of the food the user will eat.
    Macros are calculated using: (quantity_grams / 100) * value_per_100g
    
    Example:
      POST /diet/meals/1/add_item
      { "food_item_id": 42, "quantity_grams": 200 }
      
      If food #42 has 165 kcal/100g, the calculated calories = (200/100) * 165 = 330 kcal
    """
    # Verify the meal exists
    stmt = select(Meal).where(Meal.id == meal_id)
    result = await db.execute(stmt)
    meal = result.scalar_one_or_none()

    if not meal:
        raise HTTPException(
            status_code=404,
            detail=f"Meal with ID {meal_id} not found."
        )

    # Verify the food item exists
    food_stmt = select(FoodItem).where(FoodItem.id == item.food_item_id)
    food_result = await db.execute(food_stmt)
    food = food_result.scalar_one_or_none()

    if not food:
        raise HTTPException(
            status_code=404,
            detail=f"Food item with ID {item.food_item_id} not found."
        )

    # Create the meal item (the link between meal and food)
    db_item = MealItem(
        meal_id=meal_id,
        food_item_id=item.food_item_id,
        quantity_grams=item.quantity_grams,
    )
    db.add(db_item)
    await db.commit()
    await db.refresh(db_item)

    # Calculate the macros for the response
    # Formula: (quantity / 100) * value_per_100g
    qty_factor = item.quantity_grams / 100

    logger.info(
        f"Added {item.quantity_grams}g of '{food.name}' to meal ID {meal_id} "
        f"({round(qty_factor * food.calories_kcal, 2)} kcal)"
    )

    return MealItemResponse(
        id=db_item.id,
        food_item_id=food.id,
        food_item_name=food.name,
        quantity_grams=item.quantity_grams,
        calculated_calories=round(qty_factor * food.calories_kcal, 2),
        calculated_protein=round(qty_factor * food.protein_g, 2),
        calculated_carbs=round(qty_factor * food.carbs_g, 2),
        calculated_fat=round(qty_factor * food.fat_g, 2),
    )


@router.put("/plans/{plan_id}/targets", response_model=DietPlanResponse)
async def update_diet_plan_targets(
    plan_id: int,
    update: DietPlanUpdate,
    db: AsyncSession = Depends(get_db),
):
    """
    Update the macro targets of an existing diet plan.
    Only provided fields will be updated.
    """
    stmt = select(DietPlan).where(DietPlan.id == plan_id)
    result = await db.execute(stmt)
    plan = result.scalar_one_or_none()

    if not plan:
        raise HTTPException(
            status_code=404,
            detail=f"Diet plan with ID {plan_id} not found."
        )

    update_data = update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(plan, field, value)

    await db.commit()
    await db.refresh(plan)

    logger.info(f"Updated targets for plan ID {plan_id}: {update_data}")
    return plan


@router.put("/meal-items/{item_id}", response_model=MealItemResponse)
async def update_meal_item(
    item_id: int,
    update: MealItemUpdate,
    db: AsyncSession = Depends(get_db),
):
    """
    Update the quantity of a meal item.
    Returns the updated item with recalculated macros.
    """
    stmt = (
        select(MealItem)
        .where(MealItem.id == item_id)
        .options(selectinload(MealItem.food_item))
    )
    result = await db.execute(stmt)
    item = result.scalar_one_or_none()

    if not item:
        raise HTTPException(
            status_code=404,
            detail=f"Meal item with ID {item_id} not found."
        )

    item.quantity_grams = update.quantity_grams
    await db.commit()
    await db.refresh(item)

    food = item.food_item
    qty_factor = item.quantity_grams / 100

    logger.info(f"Updated meal item ID {item_id}: quantity={update.quantity_grams}g")

    return MealItemResponse(
        id=item.id,
        food_item_id=food.id,
        food_item_name=food.name,
        quantity_grams=item.quantity_grams,
        calculated_calories=round(qty_factor * food.calories_kcal, 2),
        calculated_protein=round(qty_factor * food.protein_g, 2),
        calculated_carbs=round(qty_factor * food.carbs_g, 2),
        calculated_fat=round(qty_factor * food.fat_g, 2),
    )


@router.delete("/meal-items/{item_id}", response_model=MessageResponse)
async def remove_meal_item(
    item_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    Remove a food item from a meal.
    
    This deletes the MealItem record (the link between meal and food).
    The FoodItem itself is NOT deleted — only the association.
    """
    stmt = select(MealItem).where(MealItem.id == item_id)
    result = await db.execute(stmt)
    item = result.scalar_one_or_none()

    if not item:
        raise HTTPException(
            status_code=404,
            detail=f"Meal item with ID {item_id} not found."
        )

    await db.delete(item)
    await db.commit()

    logger.info(f"Removed meal item ID {item_id}")

    return MessageResponse(
        message="Meal item removed successfully.",
        detail=f"Removed item ID {item_id} from meal."
    )


# ============================================================
# EXPORT ENDPOINTS (Excel & PDF)
# ============================================================

@router.get("/export/excel")
async def export_diet_excel(
    user_id: str = "default_user",
    db: AsyncSession = Depends(get_db),
):
    """
    Export the current diet plan (all variations) as an Excel file.
    Each variation gets its own sheet.
    """
    try:
        plan_data = await get_current_diet_full(db, user_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    meal_font = Font(bold=True, size=11)
    meal_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    target_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    variations = plan_data.get("variations", [])
    if not variations:
        # Fallback: build a single sheet from plan.meals
        variations = [{
            "name": "Principal",
            "meals": plan_data.get("meals", []),
        }]

    for var in variations:
        sheet_name = var["name"][:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14

        # Header row
        headers = ["Alimento", "Qtd (g)", "Calorias", "Proteína (g)", "Carboidratos (g)", "Gordura (g)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        row = 2
        grand_cal = grand_pro = grand_carb = grand_fat = 0.0

        for meal in var.get("meals", []):
            # Meal name row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=meal["name"])
            cell.font = meal_font
            cell.fill = meal_fill
            cell.border = thin_border
            row += 1

            meal_cal = meal_pro = meal_carb = meal_fat = 0.0

            for item in meal.get("items", []):
                ws.cell(row=row, column=1, value=item["food_item_name"]).border = thin_border
                ws.cell(row=row, column=2, value=item["quantity_grams"]).border = thin_border
                ws.cell(row=row, column=3, value=item["calculated_calories"]).border = thin_border
                ws.cell(row=row, column=4, value=item["calculated_protein"]).border = thin_border
                ws.cell(row=row, column=5, value=item["calculated_carbs"]).border = thin_border
                ws.cell(row=row, column=6, value=item["calculated_fat"]).border = thin_border

                for col in range(2, 7):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                    ws.cell(row=row, column=col).number_format = "0.0"

                meal_cal += item["calculated_calories"]
                meal_pro += item["calculated_protein"]
                meal_carb += item["calculated_carbs"]
                meal_fat += item["calculated_fat"]
                row += 1

            # Meal subtotal
            ws.cell(row=row, column=1, value=f"Subtotal {meal['name']}").font = Font(bold=True, italic=True)
            ws.cell(row=row, column=3, value=round(meal_cal, 1)).font = Font(bold=True, italic=True)
            ws.cell(row=row, column=4, value=round(meal_pro, 1)).font = Font(bold=True, italic=True)
            ws.cell(row=row, column=5, value=round(meal_carb, 1)).font = Font(bold=True, italic=True)
            ws.cell(row=row, column=6, value=round(meal_fat, 1)).font = Font(bold=True, italic=True)
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
                if col >= 2:
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                    ws.cell(row=row, column=col).number_format = "0.0"
            row += 1

            grand_cal += meal_cal
            grand_pro += meal_pro
            grand_carb += meal_carb
            grand_fat += meal_fat

        # Grand total row
        row += 1
        for col_idx, value in enumerate(["TOTAL DO DIA", "", round(grand_cal, 1), round(grand_pro, 1), round(grand_carb, 1), round(grand_fat, 1)], 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            if col_idx >= 2:
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "0.0"

        # Targets row
        row += 1
        targets = [
            "META",
            "",
            plan_data.get("target_calories", 0),
            plan_data.get("target_protein", 0),
            plan_data.get("target_carbs", 0),
            plan_data.get("target_fat", 0),
        ]
        for col_idx, value in enumerate(targets, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = total_font
            cell.fill = target_fill
            cell.border = thin_border
            if col_idx >= 2:
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "0.0"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plano_alimentar.xlsx"},
    )


@router.get("/export/pdf")
async def export_diet_pdf(
    user_id: str = "default_user",
    db: AsyncSession = Depends(get_db),
):
    """
    Export the current diet plan (all variations) as a PDF file.
    """
    try:
        plan_data = await get_current_diet_full(db, user_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=16,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Heading2"],
        fontSize=12,
        spaceAfter=4,
        textColor=colors.HexColor("#2563EB"),
    )
    variation_style = ParagraphStyle(
        "VariationTitle",
        parent=styles["Heading2"],
        fontSize=14,
        spaceBefore=12,
        spaceAfter=6,
        textColor=colors.HexColor("#1E40AF"),
    )

    elements = []

    # Title
    elements.append(Paragraph("Plano Alimentar", title_style))
    elements.append(Spacer(1, 3 * mm))

    # Targets summary
    target_data = [
        ["Meta", "Calorias", "Proteína (g)", "Carboidratos (g)", "Gordura (g)"],
        [
            "Diário",
            f"{plan_data.get('target_calories', 0):.0f}",
            f"{plan_data.get('target_protein', 0):.1f}",
            f"{plan_data.get('target_carbs', 0):.1f}",
            f"{plan_data.get('target_fat', 0):.1f}",
        ],
    ]
    target_table = RLTable(target_data, colWidths=[80, 80, 90, 100, 80])
    target_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#D1FAE5")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(target_table)
    elements.append(Spacer(1, 6 * mm))

    variations = plan_data.get("variations", [])
    if not variations:
        variations = [{"name": "Principal", "meals": plan_data.get("meals", [])}]

    for var in variations:
        elements.append(Paragraph(f"📋 {var['name']}", variation_style))

        for meal in var.get("meals", []):
            elements.append(Paragraph(meal["name"], subtitle_style))

            table_data = [["Alimento", "Qtd (g)", "Kcal", "P (g)", "C (g)", "G (g)"]]
            meal_cal = meal_pro = meal_carb = meal_fat = 0.0

            for item in meal.get("items", []):
                table_data.append([
                    item["food_item_name"],
                    f"{item['quantity_grams']:.0f}",
                    f"{item['calculated_calories']:.1f}",
                    f"{item['calculated_protein']:.1f}",
                    f"{item['calculated_carbs']:.1f}",
                    f"{item['calculated_fat']:.1f}",
                ])
                meal_cal += item["calculated_calories"]
                meal_pro += item["calculated_protein"]
                meal_carb += item["calculated_carbs"]
                meal_fat += item["calculated_fat"]

            # Subtotal row
            table_data.append([
                "Subtotal",
                "",
                f"{meal_cal:.1f}",
                f"{meal_pro:.1f}",
                f"{meal_carb:.1f}",
                f"{meal_fat:.1f}",
            ])

            col_widths = [150, 60, 60, 60, 60, 60]
            table = RLTable(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DBEAFE")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                # Subtotal row styling
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FEF3C7")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 3 * mm))

        elements.append(Spacer(1, 4 * mm))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=plano_alimentar.pdf"},
    )
